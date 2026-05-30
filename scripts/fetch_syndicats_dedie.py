"""Section dédiée Syndicats : croise BANATIC (composition, compétences)
avec OFGL (comptes financiers consolidés).

Sources :
  - BANATIC (Ministère de l'Intérieur) : ``data/banatic/intercommunalites.xlsx``
    Liste des 9 319 groupements (EPCIs + syndicats) avec, pour chacun :
      * Identité (SIREN, nom, nature juridique, dpt siège, adresse)
      * Membres (communes/EPCIs/dpts/régions)
      * 123 compétences exercées (eau, déchets, GEMAPI, urbanisme, etc.)
  - OFGL ``ofgl-base-syndicats-consolidee`` : 53 agrégats financiers par
    syndicat et par année, **consolidés** par OFGL (budget principal +
    budgets annexes - flux croisés). 1 ligne = 1 (siren × agrégat × exer).

Historique : avant 2026-05, le pipeline utilisait ``ofgl-base-syndicats``
avec un filtre ``type_de_budget = "Budget principal"`` qui ignorait
entièrement les budgets annexes (perte significative pour les syndicats
Transports/Eau/Déchets). Le passage à la base consolidée OFGL restaure
les BA tout en neutralisant les flux croisés BP↔BA (méthode OFGL
officielle, voir « Méthode de calcul des agrégats financiers »).

Sorties :
  - ``data/syndicats/syndicats-2024.json`` : liste des 6 700 syndicats
    (filtrage : SIVU/SIVOM/SMF/SMO/PETR/POLEM, EPCIs déjà gérés)
    avec meta + comptes + compétences + liste des INSEE membres
  - ``data/syndicats/commune-to-syndicats.json`` : index inverse
    INSEE → [siren_syndicats]
"""

from __future__ import annotations

import argparse
import io
import json
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

# Force UTF-8 stdout pour les caractères Unicode (Windows cp1252 par défaut)
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
BANATIC_XLSX = DATA / "banatic" / "intercommunalites.xlsx"
OUT_DIR = DATA / "syndicats"
OUT_SYND = OUT_DIR / "syndicats-2024.json"
OUT_INDEX = OUT_DIR / "commune-to-syndicats.json"
# Nouveau cache pour la base consolidée (l'ancien `ofgl-by-agregat` contient
# les valeurs BP-only de l'ancien pipeline et peut être supprimé manuellement
# après validation de la migration).
OFGL_CACHE = OUT_DIR / "ofgl-by-agregat-consolidee"

# Natures juridiques à inclure comme "syndicats" (EPCIs et MDPH exclus)
SYNDICAT_NATURES = {"SIVU", "SIVOM", "SMF", "SMO", "PETR", "POLEM"}

# Colonnes BANATIC (1-based)
COL_DEP = 1
COL_ARRDT = 2
COL_COM_SIEGE = 3       # "INSEE_SIREN - Nom" (commune siège)
COL_SIREN_GRP = 4
COL_NOM_GRP = 5
COL_NATURE = 6
COL_MODE_FIN = 7
COL_DATE_CREATION = 10
COL_TEOM = 31
COL_REOM = 32
COL_DGF = 33
COL_DOT_COMP = 34
COL_DOT_INTERCO = 35
COL_DOT_TOUR = 36
COL_POP_DGF = 37
COL_DGF_PAR_HAB = 38
COL_POT_FISCAL = 39
COL_POP_TOTALE = 40
COL_DENSITE = 41
COL_PRES_NOM = 43
COL_PRES_PRENOM = 44
COL_NB_MEMBRES = 45
COL_NB_DELEGUES = 46
COL_SIREN_MEMBRE = 47
COL_NOM_MEMBRE = 48
COL_CATEG_MEMBRE = 49
COL_POP_MEMBRE = 50
COL_NB_COMP = 51
COL_FIRST_COMP = 52
COL_LAST_COMP = 174

# Mapping SIREN commune → INSEE via la colonne "Commune siège" qui contient
# "{SIREN_commune} - {Nom}". Mais pour les membres c'est la colonne 47.
# Le SIREN commune ne donne pas directement l'INSEE — il faut une jointure.

ANNEES_SYND = [2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024]
# Base CONSOLIDÉE (BP + BA - flux croisés). Couverture 2017-2024.
OFGL_EXPORT = "https://data.ofgl.fr/api/explore/v2.1/catalog/datasets/ofgl-base-syndicats-consolidee/exports/json"
OFGL_RECORDS = "https://data.ofgl.fr/api/explore/v2.1/catalog/datasets/ofgl-base-syndicats-consolidee/records"


def _parse_compose_field(cell: str | None) -> tuple[str, str] | None:
    """Parse une cellule du type ``"{code} - {libellé}"`` en (code, libellé)."""
    if not cell:
        return None
    s = str(cell).strip()
    if " - " in s:
        a, b = s.split(" - ", 1)
        return a.strip(), b.strip()
    return s, ""


def load_banatic() -> dict[str, dict]:
    """Charge BANATIC et construit l'index ``{siren_grp: {...meta, members:[...]}}``.

    Cache JSON dans ``data/syndicats/banatic-parsed.json`` pour éviter
    de re-parser le XLSX 74 Mo (~7 min) à chaque exécution."""
    cache = OUT_DIR / "banatic-parsed.json"
    if cache.exists():
        print(f"[banatic] cache trouvé : {cache.name}")
        return json.loads(cache.read_text(encoding="utf-8"))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"[banatic] Parsing {BANATIC_XLSX.name}… (cela peut prendre 5-10 min)")
    t0 = time.time()
    wb = openpyxl.load_workbook(BANATIC_XLSX, read_only=True)
    ws = wb["Sheet1"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    # Récupère les libellés de compétences pour produire un mapping lisible
    competence_labels = [(i, header[i]) for i in range(COL_FIRST_COMP - 1, COL_LAST_COMP)]

    groupes: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        siren_grp = row[COL_SIREN_GRP - 1]
        nature = row[COL_NATURE - 1]
        if not siren_grp or nature not in SYNDICAT_NATURES:
            continue
        siren_grp_str = str(siren_grp).strip()

        # 1ère apparition : initialiser méta + compétences
        if siren_grp_str not in groupes:
            dep_code = _parse_compose_field(row[COL_DEP - 1])
            commune_siege = _parse_compose_field(row[COL_COM_SIEGE - 1])
            competences = []
            for idx, label in competence_labels:
                v = row[idx]
                if v and str(v).strip().upper() == "OUI":
                    competences.append(label)
            groupes[siren_grp_str] = {
                "siren": siren_grp_str,
                "nom": row[COL_NOM_GRP - 1] or "",
                "nature": nature,
                "dep_code": dep_code[0].split(" ")[0] if dep_code else None,
                "dep_name": dep_code[1] if dep_code else None,
                "commune_siege_siren": commune_siege[0] if commune_siege else None,
                "commune_siege_nom": commune_siege[1] if commune_siege else None,
                "mode_financement": row[COL_MODE_FIN - 1],
                "date_creation": str(row[COL_DATE_CREATION - 1] or "")[:10],
                "teom": row[COL_TEOM - 1] == "OUI",
                "reom": row[COL_REOM - 1] == "OUI",
                "population_dgf": row[COL_POP_DGF - 1],
                "population_totale": row[COL_POP_TOTALE - 1],
                "potentiel_fiscal": row[COL_POT_FISCAL - 1],
                "densite": row[COL_DENSITE - 1],
                "nb_membres_declare": row[COL_NB_MEMBRES - 1],
                "competences": competences,
                "members": [],  # rempli au fil des lignes
            }

        # Ajouter le membre courant
        siren_membre = row[COL_SIREN_MEMBRE - 1]
        if siren_membre:
            groupes[siren_grp_str]["members"].append({
                "siren": str(siren_membre).strip(),
                "nom": row[COL_NOM_MEMBRE - 1] or "",
                "categ": row[COL_CATEG_MEMBRE - 1] or "",
                "population": row[COL_POP_MEMBRE - 1],
            })

    print(f"[banatic] {len(groupes)} syndicats chargés en {time.time()-t0:.1f}s")
    # Sauve dans le cache pour éviter de re-parser
    cache.write_text(
        json.dumps(groupes, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"[banatic] cache écrit : {cache.name} ({cache.stat().st_size/1024/1024:.1f} Mo)")
    return groupes


def load_communes_meta() -> dict[str, str]:
    """Mapping SIREN commune → INSEE commune (via meta-communes-2024.json)."""
    meta_path = DATA / "communes" / "meta-communes-2024.json"
    if not meta_path.exists():
        return {}
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    out = {}
    for entry in meta.get("communes", []):
        nom = entry[0] if len(entry) > 0 else None
        insee = entry[1] if len(entry) > 1 else None
        # La synthèse-communes contient aussi le siren. Mais pas dans meta.
        # On reconstruit via la synthèse pour avoir le SIREN.
        pass
    return out


def load_siren_to_insee_commune() -> dict[str, str]:
    """Mapping SIREN commune → INSEE via synthese-communes."""
    synth = DATA / "communes" / "synthese-communes-2024.json"
    if not synth.exists():
        return {}
    d = json.loads(synth.read_text(encoding="utf-8"))
    out = {}
    for c in d.get("communes", []):
        siren = c.get("siren")
        insee = c.get("insee")
        if siren and insee:
            out[str(siren).strip()] = str(insee).strip()
    return out


def load_meta_communes_maps() -> tuple[dict[str, list[str]], dict[str, dict]]:
    """Index pour la résolution des syndicats de second degré (membres EPCI).

    Retourne deux mappings construits depuis ``meta-communes-2024.json`` :
      - ``epci_to_communes`` : ``{siren_epci: [insee, ...]}`` — permet
        d'expandre un membre ``groupement`` (EPCI) vers ses communes.
      - ``insee_meta`` : ``{insee: {"nom":..., "population":...}}`` — pour
        nommer/peupler les communes ajoutées par expansion.

    Positions meta : ``[nom, insee, dep_code, dep_name, population,
    siren_epci, siren_ept]``.
    """
    meta_path = DATA / "communes" / "meta-communes-2024.json"
    if not meta_path.exists():
        return {}, {}
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    epci_to_communes: dict[str, list[str]] = defaultdict(list)
    insee_meta: dict[str, dict] = {}
    for e in meta.get("communes", []):
        nom = e[0] if len(e) > 0 else None
        insee = str(e[1]).strip() if len(e) > 1 and e[1] else None
        population = e[4] if len(e) > 4 else None
        siren_epci = str(e[5]).strip() if len(e) > 5 and e[5] else None
        if not insee:
            continue
        insee_meta[insee] = {"nom": nom or "", "population": population}
        if siren_epci:
            epci_to_communes[siren_epci].append(insee)
    return dict(epci_to_communes), insee_meta


def download_ofgl_syndicats(force: bool = False) -> None:
    """Télécharge les comptes OFGL pour les syndicats (filtre catégorie ≠ MDPH).
    1 fichier par agrégat × année."""
    OFGL_CACHE.mkdir(parents=True, exist_ok=True)
    # Lister les agrégats (cache distinct de l'ancien `agregats_synd.json`
    # car la base consolidée peut différer en libellés/contenu)
    cache_ag = OUT_DIR / "agregats_synd_conso.json"
    if cache_ag.exists():
        agregats = json.loads(cache_ag.read_text(encoding="utf-8"))
    else:
        params = {"select": "agregat", "group_by": "agregat", "order_by": "agregat", "limit": "100"}
        url = f"{OFGL_RECORDS}?" + urllib.parse.urlencode(params)
        with urllib.request.urlopen(url, timeout=60) as r:
            d = json.loads(r.read())
        agregats = [r.get("agregat") for r in d.get("results", []) if r.get("agregat")]
        cache_ag.write_text(json.dumps(agregats, ensure_ascii=False), encoding="utf-8")
    print(f"[ofgl-synd] {len(agregats)} agrégats à télécharger")

    # On télécharge TOUTES les catégories (y compris MDPH).
    # ODSQL ne supporte pas `<>` ni `!=` proprement. Le filtrage par
    # nature juridique se fait au moment du load_ofgl_accounts en
    # croisant avec les SIREN du dict `groupes` BANATIC (qui n'inclut
    # PAS les MDPH puisqu'on filtre par natures juridiques amont).
    #
    # NB : la base consolidée n'a pas de champ `type_de_budget` (1 ligne
    # par siren×agrégat×exer, déjà consolidée BP+BA-flux). Le `montant`
    # téléchargé est directement la valeur consolidée OFGL.
    for i, ag in enumerate(agregats, 1):
        slug = ag.lower().replace(" ", "_").replace("'", "").replace("é", "e").replace("è", "e").replace("'", "")
        slug = "".join(c if c.isalnum() or c == "_" else "_" for c in slug)
        slug = slug.replace("__", "_").strip("_")[:80]
        out = OFGL_CACHE / f"{slug}.json"
        if out.exists() and not force:
            continue
        t0 = time.time()
        params = {
            "where": f'agregat = "{ag}"',
            "select": "exer,siren,montant",
        }
        url = f"{OFGL_EXPORT}?" + urllib.parse.urlencode(params)
        try:
            with urllib.request.urlopen(url, timeout=600) as r:
                out.write_bytes(r.read())
        except Exception as e:
            print(f"  ERR {ag}: {e}")
            continue
        sz = out.stat().st_size
        print(f"[ofgl-synd] [{i}/{len(agregats)}] {ag[:55]:55} -> {sz/1024:.0f} Ko en {time.time()-t0:.1f}s")


def load_ofgl_accounts() -> dict[str, dict[str, list]]:
    """Charge tous les agrégats du cache : ``{siren: {agregat: [v2017..v2024]}}``."""
    if not OFGL_CACHE.exists():
        return {}
    cache_ag = OUT_DIR / "agregats_synd_conso.json"
    if not cache_ag.exists():
        return {}
    agregats = json.loads(cache_ag.read_text(encoding="utf-8"))
    accounts: dict[str, dict[str, list]] = {}
    null_serie = [None] * len(ANNEES_SYND)
    for ag in agregats:
        slug = ag.lower().replace(" ", "_").replace("'", "").replace("é", "e").replace("è", "e").replace("'", "")
        slug = "".join(c if c.isalnum() or c == "_" else "_" for c in slug)
        slug = slug.replace("__", "_").strip("_")[:80]
        f = OFGL_CACHE / f"{slug}.json"
        if not f.exists():
            continue
        try:
            records = json.loads(f.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        for r in records:
            siren = str(r.get("siren") or "").strip()
            if not siren:
                continue
            exer = r.get("exer")
            try:
                annee = int(str(exer).split("-")[0]) if exer else None
            except ValueError:
                annee = None
            if annee is None or annee not in ANNEES_SYND:
                continue
            valeur = r.get("montant")
            if valeur is None:
                continue
            year_idx = ANNEES_SYND.index(annee)
            entry = accounts.setdefault(siren, {})
            serie = entry.setdefault(ag, list(null_serie))
            if serie[year_idx] is None:
                serie[year_idx] = float(valeur)
            else:
                serie[year_idx] += float(valeur)
    return accounts


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--skip-download", action="store_true")
    parser.add_argument("--force-download", action="store_true")
    args = parser.parse_args()

    t0 = time.time()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("=" * 60)
    print("Syndicats — Section dédiée (BANATIC + OFGL)")
    print("=" * 60)

    # 1. BANATIC
    if not BANATIC_XLSX.exists():
        print(f"ERR : {BANATIC_XLSX} manquant. Télécharge d'abord BANATIC.")
        return
    groupes = load_banatic()
    print()

    # 2. SIREN commune → INSEE
    siren_to_insee = load_siren_to_insee_commune()
    print(f"[map] {len(siren_to_insee)} mappings SIREN commune → INSEE")
    # SIREN EPCI → communes (pour expandre les syndicats de second degré dont
    # les membres sont des EPCI, cf. SI du Pays de Maurienne).
    epci_to_communes, insee_meta = load_meta_communes_maps()
    print(f"[map] {len(epci_to_communes)} EPCI → communes, {len(insee_meta)} INSEE référencés")
    print()

    # 3. OFGL syndicats : download + load
    if not args.skip_download:
        download_ofgl_syndicats(force=args.force_download)
        print()
    accounts = load_ofgl_accounts()
    print(f"[ofgl-synd] comptes chargés pour {len(accounts)} SIREN syndicats")
    print()

    # 4. Construction du JSON final
    print("Construction de syndicats-2024.json…")
    syndicats_payload = []
    n_via_epci = 0          # communes ajoutées par expansion EPCI
    n_synd_2nd_degre = 0    # syndicats ayant ≥1 membre EPCI expansé
    for siren, info in groupes.items():
        # `members` = communes résolues : membres directs (SIREN commune →
        # INSEE) ∪ communes des EPCI membres (expansion second degré),
        # dédupliquées par INSEE (le membre direct prime sur la voie EPCI).
        # `member_groups` = membres non-communes (EPCI, personne morale)
        # conservés verbatim BANATIC pour afficher la structure réelle du
        # syndicat dans le panneau détail.
        members_by_insee: dict[str, dict] = {}
        member_groups: list[dict] = []
        had_epci_expansion = False
        for m in info["members"]:
            categ = (m.get("categ") or "").strip().lower()
            insee = siren_to_insee.get(m["siren"])
            if insee:
                # Membre commune direct
                members_by_insee[insee] = {
                    "siren": m["siren"],
                    "insee": insee,
                    "nom": m["nom"],
                    "categ": m["categ"],
                    "population": m["population"],
                }
                continue
            if categ == "groupement":
                # Syndicat de second degré : expandre l'EPCI vers ses communes
                # (territoire couvert = communes de l'EPCI membre).
                communes = epci_to_communes.get(str(m["siren"]).strip(), [])
                if communes:
                    had_epci_expansion = True
                member_groups.append({
                    "siren": m["siren"],
                    "nom": m["nom"],
                    "categ": m["categ"],
                    "population": m["population"],
                    "nb_communes": len(communes),
                })
                for ins in communes:
                    if ins in members_by_insee:
                        continue  # déjà membre direct → ne pas dupliquer
                    cm = insee_meta.get(ins, {})
                    members_by_insee[ins] = {
                        "siren": None,
                        "insee": ins,
                        "nom": cm.get("nom") or "",
                        "categ": "commune (via EPCI)",
                        "population": cm.get("population"),
                        "via_epci": m["siren"],
                        "via_epci_nom": m["nom"],
                    }
                    n_via_epci += 1
            else:
                # personne morale ou membre non résoluble : verbatim, non
                # expansé (pas de communes rattachables → reste gris).
                member_groups.append({
                    "siren": m.get("siren"),
                    "nom": m.get("nom"),
                    "categ": m.get("categ"),
                    "population": m.get("population"),
                })
        if had_epci_expansion:
            n_synd_2nd_degre += 1
        members_insee = list(members_by_insee.values())
        # Comptes financiers
        comptes = accounts.get(siren, {})
        syndicats_payload.append({
            **info,
            "members": members_insee,
            "member_groups": member_groups,
            "comptes": comptes,
        })

    payload = {
        "years": ANNEES_SYND,
        "n_syndicats": len(syndicats_payload),
        "syndicats": syndicats_payload,
    }
    OUT_SYND.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    sz = OUT_SYND.stat().st_size / 1024 / 1024
    print(f"  -> {OUT_SYND.name} ({len(syndicats_payload)} syndicats, {sz:.1f} Mo)")

    # 5. Index inverse INSEE → [siren_synds]
    inv: dict[str, list[str]] = defaultdict(list)
    for s in syndicats_payload:
        for m in s["members"]:
            if m["insee"]:
                inv[m["insee"]].append(s["siren"])
    inv_dict = {k: v for k, v in inv.items()}
    OUT_INDEX.write_text(
        json.dumps(inv_dict, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"  -> {OUT_INDEX.name} ({len(inv_dict)} INSEE couverts, "
          f"{OUT_INDEX.stat().st_size/1024/1024:.1f} Mo)")

    # 6. Stats
    n_with_accounts = sum(1 for s in syndicats_payload if s["comptes"])
    print(f"\nStats finales :")
    print(f"  Syndicats avec comptes OFGL : {n_with_accounts}/{len(syndicats_payload)}")
    avg_members = sum(len(s["members"]) for s in syndicats_payload) / len(syndicats_payload)
    print(f"  Moyenne communes/syndicat : {avg_members:.1f}")
    n_synd_avec_communes = sum(1 for s in syndicats_payload if s["members"])
    print(f"  Syndicats avec ≥1 commune résolue : {n_synd_avec_communes}/{len(syndicats_payload)}")
    print(f"  Second degré (≥1 EPCI expansé) : {n_synd_2nd_degre} syndicats, "
          f"{n_via_epci} communes ajoutées via EPCI")
    n_synd_groups = sum(1 for s in syndicats_payload if s.get("member_groups"))
    print(f"  Syndicats avec member_groups (EPCI/PM) : {n_synd_groups}")

    print(f"\nTerminé en {time.time()-t0:.1f}s")


if __name__ == "__main__":
    main()
